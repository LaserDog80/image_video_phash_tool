"""Export rename results to an Excel workbook.

Produces a ``.xlsx`` file with one row per copied clip, pre-filling columns
that can be derived from the filenames and leaving blanks for user input.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from media_pairing.file_renamer import RenameResult

# Regex to extract scene number from filenames like SPT25_SC38_000601_001_V.mp4
_SCENE_RE = re.compile(r"SC(\d+)", re.IGNORECASE)

HEADERS = [
    "DESCRIPTION",
    "SCENE",
    "AI working file name",
    "AI complete file name",
    "Creation day",
]


def _parse_scene(filename: str) -> str:
    """Extract the scene number from a filename, or return empty string."""
    m = _SCENE_RE.search(filename)
    return m.group(1) if m else ""


def export_rename_to_excel(
    rename_result: RenameResult,
    output_path: str | Path,
) -> Path:
    """Write rename results to an Excel workbook.

    Columns:
        - **DESCRIPTION** — blank (user fills in)
        - **SCENE** — parsed from the output filename (``SC38`` → ``38``)
        - **AI working file name** — source video filename
        - **AI complete file name** — renamed output filename
        - **Creation day** — blank (user fills in)

    Args:
        rename_result: The result from :meth:`MediaFileRenamer.execute`.
        output_path: Where to save the ``.xlsx`` file.

    Returns:
        The resolved :class:`Path` of the saved file.
    """
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Clip Report"

    # Header row (bold)
    bold = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Data rows
    for row_idx, entry in enumerate(rename_result.copied_files, start=2):
        src_name = Path(entry["source"]).name
        dst_name = Path(entry["destination"]).name
        scene = _parse_scene(dst_name)

        ws.cell(row=row_idx, column=1, value="")        # DESCRIPTION
        ws.cell(row=row_idx, column=2, value=scene)      # SCENE
        ws.cell(row=row_idx, column=3, value=src_name)   # AI working file name
        ws.cell(row=row_idx, column=4, value=dst_name)   # AI complete file name
        ws.cell(row=row_idx, column=5, value="")         # Creation day

    # Auto-size columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(output_path)
    return output_path
