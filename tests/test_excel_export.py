"""Tests for the Excel export module."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from media_pairing.excel_export import HEADERS, _parse_scene, export_rename_to_excel
from media_pairing.file_renamer import RenameResult


# ------------------------------------------------------------------
# Scene parsing
# ------------------------------------------------------------------


class TestParseScene:
    def test_standard_scene(self):
        assert _parse_scene("SPT25_SC38_000601_001_V.mp4") == "38"

    def test_single_digit(self):
        assert _parse_scene("SPT25_SC2_001_V.mp4") == "2"

    def test_no_scene(self):
        assert _parse_scene("random_clip_001.mp4") == ""

    def test_case_insensitive(self):
        assert _parse_scene("SPT25_sc11_001_V.mp4") == "11"

    def test_multiple_sc_returns_first(self):
        assert _parse_scene("SC01_SC99_clip.mp4") == "01"


# ------------------------------------------------------------------
# Excel export
# ------------------------------------------------------------------


class TestExportRenameToExcel:
    @pytest.fixture()
    def rename_result(self) -> RenameResult:
        result = RenameResult()
        result.copied_files = [
            {
                "source": "/input/clip_mwrite_001.mp4",
                "destination": "/output/SPT25_SC38_000601_001_V.mp4",
                "type": "video",
            },
            {
                "source": "/input/clip_mwrite_002.mp4",
                "destination": "/output/SPT25_SC38_000601_002_V.mp4",
                "type": "video",
            },
            {
                "source": "/input/other_clip.mp4",
                "destination": "/output/SPT25_SC11_00101_001_V.mp4",
                "type": "video",
            },
        ]
        return result

    def test_creates_file(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        result_path = export_rename_to_excel(rename_result, out)
        assert result_path.exists()

    def test_correct_headers(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(rename_result, out)
        wb = load_workbook(out)
        ws = wb.active
        row1 = [cell.value for cell in ws[1]]
        assert row1 == HEADERS

    def test_headers_are_bold(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(rename_result, out)
        wb = load_workbook(out)
        ws = wb.active
        for cell in ws[1]:
            assert cell.font.bold

    def test_row_count(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(rename_result, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.max_row == 4  # 1 header + 3 data rows

    def test_scene_parsed(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(rename_result, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "38"
        assert ws.cell(row=4, column=2).value == "11"

    def test_source_filename(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(rename_result, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "clip_mwrite_001.mp4"

    def test_destination_filename(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(rename_result, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=4).value == "SPT25_SC38_000601_001_V.mp4"

    def test_blank_columns(self, tmp_path: Path, rename_result: RenameResult):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(rename_result, out)
        wb = load_workbook(out)
        ws = wb.active
        for row_idx in range(2, 5):
            assert not ws.cell(row=row_idx, column=1).value  # DESCRIPTION
            assert not ws.cell(row=row_idx, column=5).value  # Creation day

    def test_empty_result(self, tmp_path: Path):
        out = tmp_path / "report.xlsx"
        export_rename_to_excel(RenameResult(), out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.max_row == 1  # header only
